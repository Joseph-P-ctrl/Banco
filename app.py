from flask import Flask, jsonify, render_template, request, send_file, session, redirect, url_for
from AccountService import AccountService
from InterbankService import InterbankService
from ProviderService import ProviderService
from TransferService import TransferService
from BaseDatosService import BaseDatosService
from AsientoService import AsientoService
from io import BytesIO
from flask_session import Session
from flask_caching import Cache
from openpyxl import load_workbook
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook
import re
import logging
import traceback
import json
import smtplib
from email.message import EmailMessage
from storage_paths import ensure_data_dirs, bootstrap_bd_from_source, files_path, logs_path, SESSION_DIR, bd_path

# setup logging
ensure_data_dirs()
bootstrap_bd_from_source()
logging.basicConfig(filename=logs_path('error.log'), level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s')
app = Flask(__name__)
app.secret_key = 'AldoAbril1978'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = SESSION_DIR
Session(app)
cache = Cache(app, config={'CACHE_TYPE': 'simple'})

TRANSFER = "TRANSFER"
PROVIDERS = "PROVIDER"
INTERBANK = "INTERBAN"
CUENTA = "MOVIMIENT"
MOVIMIENTOS = "MOVIMIENTOS"
ASIENTO= "EXPORT"

def extract_emails_from_df(df):
    emails = set()
    email_regex = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    for col in df.columns:
        for val in df[col].dropna():
            for m in email_regex.findall(str(val)):
                emails.add(m)
    return sorted(emails)

def extract_emails_without_voucher(df):
    if df is None or len(df) == 0:
        return []

    candidate_cols = []
    for col in df.columns:
        normalized = str(col).strip().lower().replace('°', 'º')
        if normalized in ['nº documento', 'nºdocumento', 'asientos', 'voucher contable']:
            candidate_cols.append(col)
        elif 'documento' in normalized or 'voucher' in normalized or 'asiento' in normalized:
            candidate_cols.append(col)

    if candidate_cols:
        voucher_col = candidate_cols[0]
        voucher_values = df[voucher_col]
        empty_mask = voucher_values.isna() | (voucher_values.astype(str).str.strip() == '')
        filtered_df = df.loc[empty_mask]
    else:
        filtered_df = df

    return extract_emails_from_df(filtered_df)

def normalize_reference(value):
    return str(value).strip().upper()

def extract_single_email(value):
    if pd.isna(value):
        return ''
    matches = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", str(value))
    if matches:
        return matches[0].strip().lower()
    return ''

def build_clientes_email_map(df_clientes):
    if df_clientes is None or len(df_clientes) == 0:
        return {}

    normalized_columns = [str(c).strip().lower() for c in df_clientes.columns]
    if 'referencia' not in normalized_columns or 'correo de contacto' not in normalized_columns:
        return {}

    referencia_col = df_clientes.columns[normalized_columns.index('referencia')]
    correo_col = df_clientes.columns[normalized_columns.index('correo de contacto')]

    df_map = df_clientes[[referencia_col, correo_col]].copy()
    df_map['referencia_key'] = df_map[referencia_col].apply(normalize_reference)
    df_map['correo_clean'] = df_map[correo_col].apply(extract_single_email)
    df_map = df_map[df_map['correo_clean'] != '']

    return dict(zip(df_map['referencia_key'], df_map['correo_clean']))

def load_clientes_email_map_from_bd():
    try:
        config_path = bd_path('config.json')
        if not os.path.exists(config_path):
            return {}
        import json
        with open(config_path, 'r', encoding='utf-8') as file:
            config = json.load(file)
        clientes_file_name = config.get('CLIENTES')
        if not clientes_file_name:
            return {}
        clientes_path = bd_path(clientes_file_name)
        if not os.path.exists(clientes_path):
            return {}
        df_clientes = pd.read_excel(clientes_path, header=0)
        return build_clientes_email_map(df_clientes)
    except Exception:
        return {}

def get_no_voucher_mask(df):
    if df is None or len(df) == 0:
        return pd.Series(dtype=bool)

    candidate_cols = []
    for col in df.columns:
        normalized = str(col).strip().lower().replace('°', 'º')
        if normalized in ['nº documento', 'nºdocumento', 'asientos', 'voucher contable']:
            candidate_cols.append(col)
        elif 'documento' in normalized or 'voucher' in normalized or 'asiento' in normalized:
            candidate_cols.append(col)

    if candidate_cols:
        voucher_col = candidate_cols[0]
        voucher_values = df[voucher_col]
        return voucher_values.isna() | (voucher_values.astype(str).str.strip() == '')

    return pd.Series([True] * len(df), index=df.index)

def collect_emails_without_voucher_using_clientes(df_movimientos, clientes_email_map):
    if df_movimientos is None or len(df_movimientos) == 0 or not clientes_email_map:
        return []

    no_voucher_mask = get_no_voucher_mask(df_movimientos)
    if len(no_voucher_mask) == 0:
        return []

    if 'Correo' not in df_movimientos.columns:
        df_movimientos['Correo'] = ''

    if 'Referencia' in df_movimientos.columns:
        for index, row in df_movimientos.loc[no_voucher_mask].iterrows():
            ref_key = normalize_reference(row.get('Referencia', ''))
            if ref_key in clientes_email_map:
                df_movimientos.at[index, 'Correo'] = clientes_email_map[ref_key]

    emails = set()
    for value in df_movimientos.loc[no_voucher_mask, 'Correo'].dropna():
        clean_email = extract_single_email(value)
        if clean_email:
            emails.add(clean_email)
    return sorted(emails)

def save_emails_cache(emails):
    try:
        cache_path = files_path('emails_cache.json')
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump({'emails': emails}, f, ensure_ascii=False)
    except Exception:
        pass

def load_emails_cache():
    try:
        cache_path = files_path('emails_cache.json')
        if not os.path.exists(cache_path):
            return []
        with open(cache_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        emails = payload.get('emails', [])
        if isinstance(emails, list):
            return emails
        return []
    except Exception:
        return []

def render_correos_page(emails=None, mensaje_exito=None, page=1):
    if emails is None:
        emails = []
    page_size = 50
    total = len(emails)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    page_emails = emails[start:end]
    return render_template(
        'correos.html',
        emails=emails,
        page_emails=page_emails,
        page=page,
        total_pages=total_pages,
        page_size=page_size,
        total_emails=total,
        mensaje_exito=mensaje_exito
    )

def extract_emails_from_excel_upload(file_storage):
    emails = set()
    email_regex = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    try:
        file_storage.stream.seek(0)
        workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
        file_storage.stream.seek(0)

        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        for m in email_regex.findall(str(cell.value)):
                            emails.add(m)
                    if cell.hyperlink and cell.hyperlink.target:
                        target = str(cell.hyperlink.target)
                        if target.lower().startswith('mailto:'):
                            target = target[7:]
                        target = target.split('?', 1)[0]
                        for m in email_regex.findall(target):
                            emails.add(m)
    except Exception:
        # If the file cannot be opened with openpyxl (e.g., unsupported format), fallback to df extraction only
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass
    return sorted(emails)

@app.route('/', methods=['POST','GET'])
def home():
    if request.method == 'POST':
        files = request.files.getlist('file')
        filtered_files = [x for x in files if x.filename!=""]
        if len(filtered_files) <= 1:
            return render_template('home.html', error_message= 'Debe subir por lo menos un archivo.')
        else:
            try:
                accountService = AccountService()    
                transferService = TransferService()
                interbankService = InterbankService()
                providerService = ProviderService()
                for file in files:
                    nombre =  file.filename.upper() 
                    if (nombre != ""):
                        if CUENTA in nombre:
                            accountService.process_movements(file)
                        elif TRANSFER in nombre:  
                            transferService.setMovimientos(accountService.movimientos)
                            transferService.process_transfers(file)
                        elif INTERBANK in nombre:
                            interbankService.setMovimientos(accountService.movimientos)
                            interbankService.process_interbanks(file)
                        elif PROVIDERS in nombre:
                            providerService.setMovimientos(accountService.movimientos)
                            providerService.process_providers(file)    
                        else:
                            raise Exception("Archivo no ubicado: "+nombre)    
                guardaMovimientos(accountService.movimientos)
                guardaRecaudos(accountService.recaudos)
                resumen = {"movements": accountService.error, "providers": providerService.error, "transfers": transferService.error, "interbanks": interbankService.error}
               
                return render_template("response.html", data= resumen) 
    
                #return redirect(url_for('upload'))
            except Exception as e:
                error_message = str(e)
                return render_template('home.html', error_message= error_message)
    else:
        return render_template('home.html')

def guardaMovimientos(movimientos):
    movimientos["Fecha"] = pd.to_datetime(movimientos["Fecha"], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    excel_file = BytesIO()
    movimientos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 20  
    worksheet.column_dimensions["C"].width = 30  
    worksheet.column_dimensions["K"].width = 40  
    worksheet.column_dimensions["L"].width = 40  
    worksheet.column_dimensions["M"].width = 35 
    worksheet.column_dimensions["N"].width = 40
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        header_value = col[0].value
        if header_value == 'Correo':
            worksheet.column_dimensions[col[0].column_letter].width = 42
            break
    ruta_archivo = files_path('movimientos.xlsx')
    workbook.save(ruta_archivo)

def guardaRecaudos(recaudos):
    excel_file = BytesIO()
    recaudos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 15  
    worksheet.column_dimensions["B"].width = 50  
    worksheet.column_dimensions["C"].width = 25  
    worksheet.column_dimensions["D"].width = 10  
    worksheet.column_dimensions["E"].width = 15 
    worksheet.column_dimensions["F"].width = 10 

    ruta_archivo = files_path('recaudos.xlsx')
    workbook.save(ruta_archivo)

@app.route('/basedatos', methods=['POST','GET'])
def basedatos():
    if request.method == 'POST':
        files    = request.files.getlist('file')
        
        try:
            filtered_files = [x for x in files if x.filename!=""]
                        
            if len(filtered_files) < 1:
                return render_template('base-datos.html', error_message= 'Debe subir por lo menos un archivo.')

            nombres = [f.filename.upper() for f in filtered_files]
            valid_patterns = ['RECAUDO', 'PREPAGO', 'TRABAJADOR', 'CLIENTE']
            invalid_files = [n for n in nombres if not any(pattern in n for pattern in valid_patterns)]
            if invalid_files:
                return render_template('base-datos.html', error_message='Archivo(s) no reconocido(s): ' + ', '.join(invalid_files))

            mensaje_exito = 'Base de datos actualizada correctamente.'
            
            base_datos_service = BaseDatosService()  
            base_datos_service.GuardarAchivos(files)  
            return render_template('base-datos.html',mensaje_exito=mensaje_exito)
                
        except Exception as e:
            error_message = str(e)
            return render_template('base-datos.html', error_message= error_message)

    else:
        nohay = 'Archivo subido correctamente.'
        return render_template('base-datos.html')
    
@app.route('/asiento', methods=['POST'])
def asiento_procesar():
    logging.error('asiento_procesar: start')
    files = request.files.getlist('file')
    logging.error('asiento_procesar: received %d files', len(files))
    filtered_files = [x for x in files if x.filename!=""]
    logging.error('asiento_procesar: filtered %d files', len(filtered_files))
    if len(filtered_files) <= 1:
        logging.error('asiento_procesar: not enough files, returning form')
        return render_template('asiento.html', error_message= 'Debe subir ambos archivo.')
    else:
        try:
            asientoService = AsientoService()    
            # detect files: movimientos y asientos
            movimientosfile = None
            asientosfile = None
            for file in files:
                nombre =  file.filename.upper()
                if (nombre != ""):
                    if MOVIMIENTOS in nombre or "MOVIMIENTO" in nombre:
                        movimientosfile = file
                    elif ASIENTO in nombre:
                        asientosfile = file
                    else:
                        # ignore unknown files for now
                        pass

            if movimientosfile is None or asientosfile is None:
                raise Exception('Faltan archivos requeridos: Movimientos o Asientos')

            clientes_email_map = load_clientes_email_map_from_bd()

            asientoService.conciliar(movimientosfile, asientosfile)
            #solo si hay asientos se completa en el cache
            if asientoService.df_movimientos is not None:
                emails = collect_emails_without_voucher_using_clientes(asientoService.df_movimientos, clientes_email_map)
                guardaAsientos(asientoService.df_movimientos)
                # guardar en session para uso posterior y redirigir al flujo de correos
                sorted_emails = sorted(set(emails))
                session['asiento_emails'] = sorted_emails
                save_emails_cache(sorted_emails)
                if len(sorted_emails) == 0:
                    session['asiento_email_warning'] = 'No se encontraron correos para líneas sin voucher. Verifique Referencia y CORREO DE CONTACTO en Base de Datos > Clientes.'
                else:
                    session.pop('asiento_email_warning', None)
                # guardaAsientos ya escribió files/asientos.xlsx, descargamos directamente
                ruta_archivo = files_path('asientos.xlsx')
                return send_file(ruta_archivo, as_attachment=True, download_name='asientos.xlsx')
            else: 
                #si hubiera error se pinta la misma pagina y no se redirecciona
                return render_template('asiento.html', error_message= 'No se encontro ningun asiento en el proceso')       
            
        except Exception as e:
            error_message = str(e)
            logging.error('asiento_procesar: exception: %s', error_message)
            return render_template('asiento.html', error_message= error_message)
    # Fallback: ensure the view always returns a response
    logging.error('asiento_procesar: reached end of function without explicit return')
    return render_template('asiento.html', error_message='Error inesperado en el procesamiento')



@app.route('/asiento', methods=['GET'])
def asiento_get():
    return render_template('asiento.html')


@app.route('/correos', methods=['GET','POST'])
def correos():
    sess_emails = session.get('asiento_emails', [])
    if not sess_emails:
        sess_emails = load_emails_cache()
        if sess_emails:
            session['asiento_emails'] = sess_emails
    warning_message = session.pop('asiento_email_warning', None)

    try:
        page = int(request.args.get('page', '1'))
    except ValueError:
        page = 1
    if page < 1:
        page = 1

    return render_correos_page(emails=sess_emails, mensaje_exito=warning_message, page=page)

@app.route('/upload', methods=['POST','GET'])
def upload():
    
    if request.method == 'POST':
        ruta_archivo = files_path('movimientos.xlsx')
        return send_file(ruta_archivo, as_attachment=True, download_name="movimientos.xlsx")
   
    

@app.route('/download_recaudos', methods=['POST'])
def download_recaudos():
    ruta_archivo = files_path('recaudos.xlsx')
    return send_file(ruta_archivo, as_attachment=True, download_name="recaudos.xlsx")

def guardaAsientos(movimientosAsientos):
    movimientosAsientos["Fecha"] = pd.to_datetime(movimientosAsientos["Fecha"], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    excel_file = BytesIO()
    movimientosAsientos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 20  
    worksheet.column_dimensions["C"].width = 30  
    worksheet.column_dimensions["K"].width = 40  
    worksheet.column_dimensions["L"].width = 40  
    worksheet.column_dimensions["M"].width = 35 
    worksheet.column_dimensions["N"].width = 28 

    ruta_archivo = files_path('asientos.xlsx')
    workbook.save(ruta_archivo)


@app.route('/download_asientos', methods=['POST'])
def dowload_asientos():
    ruta_archivo = files_path('asientos.xlsx')
    return send_file(ruta_archivo, as_attachment=True, download_name="Asiento.xlsx")


@app.route('/send_emails', methods=['POST'])
def send_emails():
    emails = session.get('asiento_emails', [])
    if not emails:
        return render_correos_page(emails=[], mensaje_exito='No hay correos para enviar', page=1)
    selected_emails = request.form.getlist('selected_emails')
    if selected_emails:
        emails_to_send = sorted(set(selected_emails))
    else:
        emails_to_send = emails
    sender = os.environ.get('GMAIL_SENDER', '').strip()
    app_password = os.environ.get('GMAIL_APP_PASSWORD', '').strip()
    subject = os.environ.get('GMAIL_SUBJECT', 'Notificación de Atención')
    body = os.environ.get('GMAIL_BODY', 'Estimado(a),\n\nSe comparte la notificación correspondiente.\n\nSaludos.')

    if not sender or not app_password:
        return render_correos_page(
            emails=emails,
            mensaje_exito='Falta configurar Gmail. Define GMAIL_SENDER y GMAIL_APP_PASSWORD (App Password de Gmail).',
            page=1
        )

    sent_ok = []
    sent_fail = []
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=30) as smtp:
            smtp.login(sender, app_password)
            for recipient in emails_to_send:
                try:
                    msg = EmailMessage()
                    msg['Subject'] = subject
                    msg['From'] = sender
                    msg['To'] = recipient
                    msg.set_content(body)
                    smtp.send_message(msg)
                    sent_ok.append(recipient)
                except Exception as send_ex:
                    sent_fail.append(f"{recipient}: {send_ex}")

        report_path = files_path('emails_send_report.csv')
        with open(report_path, 'w', encoding='utf-8') as report:
            report.write('estado,email,detalle\n')
            for ok in sent_ok:
                report.write(f'sent,{ok},ok\n')
            for fail in sent_fail:
                email_part = fail.split(':', 1)[0]
                detail_part = fail.split(':', 1)[1].strip() if ':' in fail else 'error'
                report.write(f'failed,{email_part},"{detail_part}"\n')

        message = f"Envío finalizado. Enviados: {len(sent_ok)}. Fallidos: {len(sent_fail)}."
        if len(sent_fail) > 0:
            message += ' Revisa emails_send_report.csv para detalles.'
        return render_correos_page(emails=emails, mensaje_exito=message, page=1)
    except Exception as ex:
        return render_correos_page(emails=emails, mensaje_exito='Error enviando por Gmail: ' + str(ex), page=1)



@app.errorhandler(Exception)
def handle_exception(e):
     # Log full traceback to file
     tb = traceback.format_exc()
     logging.error('Unhandled exception:\n%s', tb)
     # return a friendly error page
     return render_template('error.html', message=str(e)), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0')

    